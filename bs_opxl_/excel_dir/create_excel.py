import openpyxl, os
from datetime import datetime, timedelta
from date_dir.date_ import data_date_

# 获取项目根目录
PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))

# 定义统一的文件路径
EXCEL_FILE_PATH = os.path.join(PROJECT_ROOT, 'static', 'NBA.xlsx')

def create_excel(filename=EXCEL_FILE_PATH):
    """创建或更新Excel文件"""
    
    sheet_names = data_date_
    excel_header = ['类型','排名','姓名','数据',' ','类型','排名','姓名','数据',' ','类型','排名','姓名','数据']
    
    if os.path.exists(filename):
        wb = openpyxl.load_workbook(filename)
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        for i, name in enumerate(sheet_names):
            if name not in wb.sheetnames:
                stats_sheet = wb.create_sheet(title=name, index=i)
                stats_sheet.append(excel_header)
                print(f'工作表创建完成: {name}')
        wb.save(filename)
        print(f'Excel文件更新完成: {filename}')
        return wb
    else:
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        for i, name in enumerate(sheet_names):
            stats_sheet = wb.create_sheet(title=name, index=i)
            stats_sheet.append(excel_header)
            print(f'工作表创建完成: {name}')
        wb.save(filename)
        print(f'Excel文件创建完成: {filename}')
        return wb



# 创建主Excel文件
Excel_ = create_excel(EXCEL_FILE_PATH)

