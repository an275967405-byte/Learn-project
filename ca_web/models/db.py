"""
数据库操作模块 - 负责Excel文件的读写操作
"""
import openpyxl
from datetime import datetime
from config import EXCEL_PATH


class StudentDB:
    """学生数据库类"""
    
    def __init__(self):
        self.filename = EXCEL_PATH
        self.wb = None
        self.ws = None
        self.students = []
        self._init_workbook()
        self._load_students()
    
    def _init_workbook(self):
        """初始化工作簿"""
        try:
            self.wb = openpyxl.load_workbook(self.filename)
            self.ws = self.wb.active
        except FileNotFoundError:
            self.wb = openpyxl.Workbook()
            self.ws = self.wb.active
            self.ws.append(["id", "name", "score", "reg_time"])
            self.wb.save(self.filename)
    
    def _load_students(self):
        """从Excel加载学生数据到内存"""
        self.students = []
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:  # 跳过空行
                student = {
                    "id": row[0],
                    "name": row[1],
                    "score": row[2],
                    "reg_time": row[3]
                }
                self.students.append(student)
    
    def get_all_students(self):
        """获取所有学生"""
        return self.students
    
    def find_by_name(self, name):
        """根据姓名查找学生"""
        for student in self.students:
            if student["name"] == name:
                return student
        return None
    
    def find_by_id(self, student_id):
        """根据ID查找学生"""
        for student in self.students:
            if student["id"] == student_id:
                return student
        return None
    
    def add_student(self, name, score):
        """添加学生"""
        new_id = max(s["id"] for s in self.students) + 1 if self.students else 1
        reg_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # 添加到Excel
        self.ws.append([new_id, name, score, reg_time])
        self.wb.save(self.filename)
        
        # 添加到内存
        new_student = {
            "id": new_id,
            "name": name,
            "score": score,
            "reg_time": reg_time
        }
        self.students.append(new_student)
        
        return new_student
    
    def update_student(self, name, new_score):
        """更新学生成绩"""
        # 更新Excel
        for row in self.ws.iter_rows(min_row=2, values_only=False):
            if row[1].value == name:
                row[2].value = new_score
                self.wb.save(self.filename)
                
                # 更新内存
                for student in self.students:
                    if student["name"] == name:
                        student["score"] = new_score
                        return student
        return None
    
    def delete_student(self, name):
        """删除学生"""
        # 从Excel删除
        for idx, row in enumerate(self.ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[1] == name:
                self.ws.delete_rows(idx)
                self.wb.save(self.filename)
                
                # 从内存删除
                self.students = [s for s in self.students if s["name"] != name]
                return True
        return False
    
    def get_scores(self):
        """获取所有成绩"""
        return [s["score"] for s in self.students]
    
    def calculate_average(self):
        """计算平均分"""
        scores = self.get_scores()
        if scores:
            return sum(scores) / len(scores)
        return None
    
    def get_max_score(self):
        """获取最高分"""
        scores = self.get_scores()
        return max(scores) if scores else None
    
    def get_min_score(self):
        """获取最低分"""
        scores = self.get_scores()
        return min(scores) if scores else None


# 创建全局数据库实例
db = StudentDB()

