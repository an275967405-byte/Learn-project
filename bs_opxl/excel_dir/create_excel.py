import openpyxl,os
from datetime import datetime, timedelta
from date_dir.date_ import data_date_

def create_excel(filename):
    sheet_names = data_date_
    excel_header = ['类型','排名','姓名','数据',' ','类型','排名','姓名','数据',' ','类型','排名','姓名','数据']
    if os.path.exists(filename):
        wb = openpyxl.load_workbook(filename)
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        for i,name in enumerate(sheet_names):
            if name not in wb.sheetnames :
                stats_sheet = wb.create_sheet(title = name,index=i)
                stats_sheet.append(excel_header)
                print(f'complete! {name}')
        wb.save(filename)
        print('create completed !')
        return wb
    else:
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        for i,name in enumerate(sheet_names) :
            stats_sheet = wb.create_sheet(title = name,index=i)
            stats_sheet.append(excel_header)
            print(f'complete! {name}')
        wb.save(filename)
        print('create completed !')
        return wb

Excel_ = create_excel('NBA.xlsx')

